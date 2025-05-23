import openpyxl
import pandas as pd
from openpyxl import workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import shutil

#Definir formato de fecha
fecha_actual = datetime.now()
dia = fecha_actual.day
mes = fecha_actual.strftime('%b')
año = fecha_actual.year
Fecha_Mantenimiento = "{}{}{}".format(dia,mes,año)

def create_template(Nombre_Tecnico, ID_Feeder, Tipo_Feeder, Fecha_Mantenimiento, Color_Semana,Observaciones):
    '''
    Funcionamiento basico:
        - Cargar el archivo Excel existente(MF-64_plantilla.XLS)
        - lo convertimos a XLSX(MF-64_plantilla.xlsx)
        ** version mas actual de archivos excel
        -recabamos toda la info de los parametros y generamos una plantilla para guardar como reporte
    parametros:
        -Nombre del tecnico
        -ID feeder
        -Tipo de feeder
        -Fecha de mantenimiento
        -Color de semana
    '''
    nombre_excel = r'H:\Ingenieria\Ensamble PCB\Documentacion ISO-9001\MF-64_plantilla.xlsx'
    
    try:
        # Intentar cargar el archivo existente
        workbook = openpyxl.load_workbook(nombre_excel)
    except FileNotFoundError:
        # Si el archivo no existe, crear uno nuevo
        workbook = Workbook()
        workbook.save(nombre_excel)
    
    #*****************Obtener la hoja de trabajo*******************
    hoja = workbook.active
    font = Font(name='Arial', size=10, bold=True)
    alignment = Alignment(horizontal='center', vertical='center')
    alignment_left = Alignment(horizontal='left', vertical='center')
    #*******************fusion de celdas*************************
    #fecha
    hoja.merge_cells("B12:C12")
    #tenico
    hoja.merge_cells("F12:G12")
    #id feeder
    hoja.merge_cells("B14:C14")
    #color
    hoja.merge_cells("F14:G14")
    #observaciones 
    hoja.merge_cells("A65:H86")
    #iamgen
    hoja.merge_cells("C4:E5")
    
    #********************************* Imagen ***********************
    img = Image('mantto_feeder\static\img\LOGO_NAVICO_1_90-black.png')
    #hoja.add_image(img, "C4")

    
    #******************** Rellenar celdas ***********************
    hoja["B12"].value = Fecha_Mantenimiento
    hoja["F12"].value = Nombre_Tecnico
    hoja["B14"].value = ID_Feeder
    hoja["F14"].value = Color_Semana
    
    if Tipo_Feeder == "CP":
        hoja["H18"].value = "OK"
    elif Tipo_Feeder == "QP":
        hoja["H26"].value = "OK"
    elif Tipo_Feeder == "BFC":
        hoja["H32"].value = "OK"
    elif Tipo_Feeder == "HOOVER":
        hoja["H39"].value = "OK"    
    hoja["A65"].value = Observaciones
    #*********************Aplicar estilos************************
    hoja["B12"].font = font
    hoja["F12"].font = font
    hoja["B14"].font = font
    hoja["F14"].font = font
    hoja["H18"].font = font
    
    hoja["B12"].alignment = alignment
    hoja["F12"].alignment = alignment
    hoja["B14"].alignment = alignment
    hoja["F14"].alignment = alignment
    hoja["H18"].alignment = alignment
    
    #guardar workbook
    ID_feeder = hoja["B14"].value
    nuevo_nombre = (f"H:\\Ingenieria\\Ensamble PCB\\Documentacion ISO-9001\\reporte de mantto feeder\\MF-64_{ID_feeder}_{Fecha_Mantenimiento.replace('/','_')}.xlsx")
    workbook.save(nuevo_nombre)
    workbook.close()
    
    
    #return create_template.__doc_
    #print(create_template().__doc__)

#create_template(Nombre_Tecnico, ID_Feeder,Tipo_Feeder,Fecha_Mantenimiento, Color_Semana)
#create_template("Cristian Mendoza", 23760056, "QP", Fecha_Mantenimiento, "AZUL","Feeder con mantenimiento atrasado")